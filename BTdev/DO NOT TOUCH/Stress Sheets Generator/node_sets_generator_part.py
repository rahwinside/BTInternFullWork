# -*- coding: utf-8 -*-
"""
Created on Thu Jun 15 11:30:58 2017

@author: ssurabhi
"""

import openpyxl

def node_zone_seperator(nodes):
    zones = {}
    for node in nodes:
        if node[:3] not in zones:
                zones[node[:3]] = [node]
        else:
            zones[node[:3]].append(node)
    return zones

excel_input = openpyxl.load_workbook('input_data.xlsx')
jauges_sheet = excel_input['jauges']
rosettes_sheet = excel_input['rosettes']
jauges_matrix = [[str(i.value) for i in j] for j in jauges_sheet]
rosettes_matrix = [[str(i.value) for i in j] for j in rosettes_sheet]

gauge_nos = []
node_ids = ""
jauge_stress_direction = []
coord_sys = []
for j2 in range(1,len(jauges_matrix)):
    node_ids += jauges_matrix[j2][1] + ' '
    gauge_nos.append(jauges_matrix[j2][0])
    jauge_stress_direction.append(jauges_matrix[j2][3])
    coord_sys.append(jauges_matrix[j2][2])
rosette_a_direction = []
rosette_c_direction = []
for j2 in range(1,len(rosettes_matrix)):
    node_ids += rosettes_matrix[j2][1] + ' '
    gauge_nos.append(rosettes_matrix[j2][0])
    rosette_a_direction.append(rosettes_matrix[j2][3])
    rosette_c_direction.append(rosettes_matrix[j2][4])
    coord_sys.append(rosettes_matrix[j2][2])
nodes_list = node_ids.split(' ')[:-1]
 
node_gauge = {}
for j9 in range(len(nodes_list)):
    node_gauge[nodes_list[j9]] = gauge_nos[j9]

node_cord = {}
for j20 in range(len(nodes_list)):
    node_cord[nodes_list[j20]] = coord_sys[j20]

systems = {}
for j21 in node_cord:
    if node_cord[j21] in systems:
        systems[node_cord[j21]].append(j21)
    else:
        systems[node_cord[j21]] = [j21]

for j22 in systems:
    txt_file_name = 'node_ids_' + j22 + '.txt'
    G = open(txt_file_name,'w')
    line1 = '*Group "' + j22 + ' Set" "node" "255   0   0"\n'
    G.write(line1)
    G.write('*Attributes point 3\n')
    G.write('*Pool "Node"\n')
    node_pool = ''
    print (node_pool)
    print (systems[j22])
    for j23 in systems[j22]:
        node_pool += j23 + ' '
    G.write(node_pool)
    G.close()


all_values = []
for node in nodes_list:
    filename ="Nodal_Stress_Reports__Iteration12_exp_vert\\Node_" + node + '_StressReport.csv'

    temp_mat = []
    G = open(filename, 'r')
    for j3 in G:
        temp_mat.append(j3[:-1].split(','))
    G.close()
    all_values.append(temp_mat)
all_values_t = [[[all_values[k][j][i] for j in range(len(all_values[0]))] for i in range(len(all_values[0][0]))] for k in range(len(all_values))]

excel_main = openpyxl.Workbook()
sheet_main = excel_main.active
sheet_main.title = 'all_stresses'

rng = 1

cng = 2
for j4 in range(len(all_values_t)):
    for j5 in range(1, len(all_values_t[0])):
        for j6 in range(len(all_values_t[0][0])):
            rn = rng + j5
            cn = cng + j6
            sheet_main.cell(row=rn, column=cn).value = all_values_t[j4][j5][j6]
    rng += j5
# print(rng)

for j7 in range(3, len(all_values_t[0][0]) + 2):
    sheet_main.cell(row=1, column=j7).value = all_values_t[0][0][j7 - 2]



nn = -1
for j8 in range(2, len(all_values_t) * 6 + 2):
    if ((j8 - 2) % 6 == 0):
        nn += 1
        sheet_main.cell(row=j8, column=1).value = node_gauge[str(nodes_list[nn])]
        # print(nn)

all_rows_gen_cs = sheet_main.rows
all_columns_gen_cs = sheet_main.columns
all_rows_cs = [[i for i in j] for j in all_rows_gen_cs]
all_columns_cs = [[i for i in j] for j in all_columns_gen_cs]
R_cs = len(all_rows_cs)
C_cs = len(all_columns_cs)
nodes = [str(all_columns_cs[0][n].value) for n in range(1, R_cs, 6)]
# print(nodes)

J_nodes = [node for node in nodes if node[0] == 'J']
R_nodes = [node for node in nodes if node[0] == 'R']
J_nodes_zones = node_zone_seperator(J_nodes)
R_nodes_zones = node_zone_seperator(R_nodes)

stresses_compr = [[all_columns_cs[0][j].value] for j in
                  range(1, len(all_columns_cs[0]) - len(R_nodes) * 6)
                  if all_columns_cs[0][j].value != None]
zrsiga = [[all_columns_cs[0][j].value] for j in range(1, len(all_columns_cs[0])) if
          all_columns_cs[0][j].value != None]
zrsigc = [[all_columns_cs[0][j].value] for j in range(1, len(all_columns_cs[0])) if
          all_columns_cs[0][j].value != None]
z = (len(zrsiga))
for j in range(1, z, 1):
    if (str(zrsiga[j][0])[:1] == 'R'):
        rosette = zrsiga[j][0]
        zrsiga[j][0] = [rosette + ' sig a']
zrsiga.insert(0, [i.value for i in all_rows_cs[0]][1:])

for j in range(1, z, 1):
    if (str(zrsigc[j][0])[:1] == 'R'):
        rosette = zrsigc[j][0]
        zrsigc[j][0] = [rosette + ' sig c']
zrsigc.insert(0, [i.value for i in all_rows_cs[0]][1:])

for i in range(len(stresses_compr) + 1, len(stresses_compr) + len(R_nodes) + 1):
    stresses_compr.append(zrsiga[i][0])
    stresses_compr.append(zrsigc[i][0])

stresses_compr.insert(0, [i.value for i in all_rows_cs[0]][1:])
node_i = 0
rosette_directions = []
for i in range(len(rosette_a_direction)):
    rosette_directions.append(rosette_a_direction[i])
    rosette_directions.append(rosette_c_direction[i])
gauges_rosettes_a_c_indices = [i for i in range(1, R_cs - len(R_nodes) * 6)]
# for rac in range(R_cs - len(R_nodes)*6,R_cs):
#    gauges_rosettes_a_c_indices.append(rac)
#    gauges_rosettes_a_c_indices.append(rac)
for j3 in gauges_rosettes_a_c_indices:
    if ((j3 - 1) % 6 == 0):
        node_i += 1
    stress_type = [jauge_stress_direction + rosette_directions][0][node_i - 1]
    for j4 in range(2, C_cs):
        if (str(all_rows_cs[j3][1].value).strip() == stress_type.strip()):
            stresses_compr[node_i].append(all_rows_cs[j3][j4].value)

node_i = len(J_nodes) - 1
gauges_rosettes_a_c_indices = [i for i in range(R_cs - len(R_nodes) * 6, R_cs)]
for j3 in gauges_rosettes_a_c_indices:
    if ((j3 - 1) % 6 == 0):
        node_i += 2
    stress_type = [jauge_stress_direction + rosette_directions][0][node_i - 1]
    for j4 in range(2, C_cs):
        if (str(all_rows_cs[j3][1].value).strip() == stress_type.strip()):
            stresses_compr[node_i].append(all_rows_cs[j3][j4].value)

node_i = len(J_nodes)
gauges_rosettes_a_c_indices = [i for i in range(R_cs - len(R_nodes) * 6, R_cs)]
for j3 in gauges_rosettes_a_c_indices:
    if ((j3 - 1) % 6 == 0):
        node_i += 2
    stress_type = [jauge_stress_direction + rosette_directions][0][node_i - 1]
    for j4 in range(2, C_cs):
        if (str(all_rows_cs[j3][1].value).strip() == stress_type.strip()):
            stresses_compr[node_i].append(all_rows_cs[j3][j4].value)
            # print(stresses_compr)

sheet_cs_u = excel_main.create_sheet('unique_stresses_only')

for j5 in range(1, len(stresses_compr) + 1):
    for j6 in range(1, len(stresses_compr[j5 - 1]) + 1):
        sheet_cs_u.cell(row=j5, column=j6).value = stresses_compr[j5 - 1][j6 - 1]

excel_co = openpyxl.load_workbook('correlation.xlsx')
sheet_co = excel_co.active
all_rows_gen_co = sheet_co.rows
all_rows_co = [[i for i in j] for j in all_rows_gen_co]

diff_error = [[all_rows_co[j][0].value] for j in range(1, len(all_rows_co))]
diff_error.insert(0, all_rows_cs[0])

diff_mat = [
    [all_rows_co[i][j].value - float(stresses_compr[i][j]) for j in range(1, len(stresses_compr[i]))]
    for i
    in range(1, len(stresses_compr))]

sheet_cs_u_diff = excel_main.copy_worksheet(sheet_cs_u)
sheet_cs_u_diff.title = 'stresses_difference'
for j7 in range(2, len(diff_mat) + 2):
    for j8 in range(2, len(diff_mat[j7 - 2]) + 2):
        sheet_cs_u_diff.cell(row=j7, column=j8).value = diff_mat[j7 - 2][j8 - 2]
        # print((stresses_compr))

for zone in J_nodes_zones:
    sheet_zone_name = 'Zone' + zone[-1] + '00' + zone[0].lower()
    sheet_zone = excel_main.create_sheet(sheet_zone_name)
    for j9 in range(len(stresses_compr[0])):
        sheet_zone.cell(row=1, column=j9 + 1).value = stresses_compr[0][j9]

    row_n = 1
    for j10 in range(1, len(stresses_compr)):
        if str(stresses_compr[j10][0][:3]) == zone and str(stresses_compr[j10][0]) in J_nodes_zones[zone]:
            row_n += 1
            for j11 in range(len(stresses_compr[j10])):
                sheet_zone.cell(row=row_n, column=j11 + 1).value = stresses_compr[j10][j11]
            # print(stresses_compr)

    row_n += 3
    for j9 in range(len(all_rows_co[0])):
        sheet_zone.cell(row=row_n, column=j9 + 1).value = all_rows_co[0][j9].value
    for j10 in range(1, len(all_rows_co)):
        if str(all_rows_co[j10][0].value[:3]) == zone and str(all_rows_co[j10][0].value) in \
                J_nodes_zones[zone]:
            row_n += 1
            for j11 in range(len(all_rows_co[j10])):
                sheet_zone.cell(row=row_n, column=j11 + 1).value = all_rows_co[j10][j11].value

    row_n += 3
    for j9 in range(len(all_rows_co[0])):
        sheet_zone.cell(row=row_n, column=j9 + 1).value = all_rows_co[0][j9].value
    for j10 in range(1, len(diff_mat) + 1):
        if str(all_rows_co[j10][0].value[:3]) == zone and str(all_rows_co[j10][0].value) in \
                J_nodes_zones[
                    zone]:
            row_n += 1
            sheet_zone.cell(row=row_n, column=1).value = all_rows_co[j10][0].value
            for j11 in range(1, len(diff_mat[j10 - 1]) + 1):
                sheet_zone.cell(row=row_n, column=j11 + 1).value = diff_mat[j10 - 1][j11 - 1]

for zone in R_nodes_zones:
    sheet_zone_name = 'Zone' + zone[-1] + '00' + zone[0].lower()
    sheet_zone = excel_main.create_sheet(sheet_zone_name)
    for j9 in range(len(stresses_compr[0])):
        sheet_zone.cell(row=1, column=j9 + 1).value = stresses_compr[0][j9]

    row_n = 1
    for j10 in range(1, len(stresses_compr)):
        if str(stresses_compr[j10][0][:3]) == zone and str(stresses_compr[j10][0][:5]) in R_nodes_zones[
            zone]:
            row_n += 1
            for j11 in range(len(stresses_compr[j10])):
                sheet_zone.cell(row=row_n, column=j11 + 1).value = stresses_compr[j10][j11]

    row_n += 3
    for j9 in range(len(all_rows_co[0])):
        sheet_zone.cell(row=row_n, column=j9 + 1).value = all_rows_co[0][j9].value
    for j10 in range(1, len(all_rows_co)):
        if str(all_rows_co[j10][0].value[:3]) == zone and str(all_rows_co[j10][0].value[:5]) in \
                R_nodes_zones[zone]:
            row_n += 1
            for j11 in range(len(all_rows_co[j10])):
                sheet_zone.cell(row=row_n, column=j11 + 1).value = all_rows_co[j10][j11].value

    row_n += 3
    for j9 in range(len(all_rows_co[0])):
        sheet_zone.cell(row=row_n, column=j9 + 1).value = all_rows_co[0][j9].value
    for j10 in range(1, len(diff_mat) + 1):
        if str(all_rows_co[j10][0].value[:3]) == zone and str(all_rows_co[j10][0].value[:5]) in \
                R_nodes_zones[zone]:
            row_n += 1
            sheet_zone.cell(row=row_n, column=1).value = all_rows_co[j10][0].value
            for j11 in range(1, len(diff_mat[j10 - 1]) + 1):
                sheet_zone.cell(row=row_n, column=j11 + 1).value = diff_mat[j10 - 1][j11 - 1]

excel_main.save("stress_sheets_final.xlsx")

