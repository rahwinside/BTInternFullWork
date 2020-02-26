import openpyxl


def node_zone_seperator(nodes):
    zones = {}
    for node in nodes:
        if node[:3] not in zones:
            zones[node[:3]] = [node]
        else:
            zones[node[:3]].append(node)
            # print(zones)
    return zones


excel_input = openpyxl.load_workbook('input_data.xlsx')
jauges_sheet = excel_input['jauges']
rosettes_sheet = excel_input['rosettes']
jauges_matrix = [[str(i.value) for i in j] for j in jauges_sheet]
rosettes_matrix = [[str(i.value) for i in j] for j in rosettes_sheet]
# print(rosettes_matrix)
gauge_nos = []
node_ids = ""
jauge_stress_direction = []
coord_sys = []
for j2 in range(1, len(jauges_matrix)):
    node_ids += jauges_matrix[j2][1] + ' '
    gauge_nos.append(jauges_matrix[j2][0])
    jauge_stress_direction.append(jauges_matrix[j2][3])
    coord_sys.append(jauges_matrix[j2][2])
rosette_a_direction = []
rosette_c_direction = []
for j2 in range(1, len(rosettes_matrix)):
    node_ids += rosettes_matrix[j2][1] + ' '
    gauge_nos.append(rosettes_matrix[j2][0])
    rosette_a_direction.append(rosettes_matrix[j2][3])
    rosette_c_direction.append(rosettes_matrix[j2][4])
    coord_sys.append(rosettes_matrix[j2][2])
nodes_list = node_ids.split(' ')[:-1]

node_gauge = {}
for j9 in range(len(nodes_list)):
    node_gauge[nodes_list[j9]] = gauge_nos[j9]
    all_values = []
for node in nodes_list:
                filename = "Nodal_Stress_Reports\\Node_" + node + '_StressReport.csv'

                temp_mat = []
                G = open(filename, 'r')
                for j3 in G:
                    temp_mat.append(j3[:-1].split(','))
                G.close()
                all_values.append(temp_mat)
            all_values_t = [
                [[all_values[k][j][i] for j in range(len(all_values[0]))] for i in range(len(all_values[0][0]))] for k
                in range(len(all_values))]

excel_main = openpyxl.Workbook()
sheet_main = excel_main.active
sheet_main.title = 'all_stresses'

rng = 1

cng = 2
for j4 in range(len(all_values_t)):
                for j5 in range(1, len(all_values_t[0])):
                    for j6 in range(len(all_values_t[0][0])):
                        rn = rng + j5
                        cn = cng + j6
                        sheet_main.cell(row=rn, column=cn).value = all_values_t[j4][j5][j6]
                rng += j5


            for j7 in range(3, len(all_values_t[0][0]) + 2):
                sheet_main.cell(row=1, column=j7).value = all_values_t[0][0][j7 - 2]

            nn = -1
            for j8 in range(2, len(all_values_t) * 6 + 2):
                if ((j8 - 2) % 6 == 0):
                    nn += 1
                    sheet_main.cell(row=j8, column=1).value = node_gauge[str(nodes_list[nn])]
                    # print(nn)

            all_rows_gen_cs = sheet_main.rows
            all_columns_gen_cs = sheet_main.columns
            all_rows_cs = [[i for i in j] for j in all_rows_gen_cs]
            all_columns_cs = [[i for i in j] for j in all_columns_gen_cs]
            R_cs = len(all_rows_cs)
            C_cs = len(all_columns_cs)
            nodes = [str(all_columns_cs[0][n].value) for n in range(1, R_cs, 6)]
            #print(C_cs)

            J_nodes = [node for node in nodes if node[0] == 'J']
            R_nodes = [node for node in nodes if node[0] == 'R']
            J_nodes_zones = node_zone_seperator(J_nodes)
            R_nodes_zones = node_zone_seperator(R_nodes)


            stresses_compr = [[all_columns_cs[0][j].value] for j in
                              range(1, len(all_columns_cs[0]) - len(R_nodes) * 6)
                              if all_columns_cs[0][j].value != None]
            zrsiga = [[all_columns_cs[0][j].value] for j in range(1, len(all_columns_cs[0])) if
                      all_columns_cs[0][j].value != None]
            zrsigc = [[all_columns_cs[0][j].value] for j in range(1, len(all_columns_cs[0])) if
                      all_columns_cs[0][j].value != None]
            z = (len(zrsiga))
            for j in range(1, z, 1):
                if (str(zrsiga[j][0])[:1] == 'R'):
                    rosette = zrsiga[j][0]
                    zrsiga[j][0] = [rosette + ' sig a']
            zrsiga.insert(0, [i.value for i in all_rows_cs[0]][1:])

            for j in range(1, z, 1):
                if (str(zrsigc[j][0])[:1] == 'R'):
                    rosette = zrsigc[j][0]
                    zrsigc[j][0] = [rosette + ' sig c']
            zrsigc.insert(0, [i.value for i in all_rows_cs[0]][1:])

            for i in range(len(stresses_compr) + 1, len(stresses_compr) + len(R_nodes) + 1):
                stresses_compr.append(zrsiga[i][0])
                stresses_compr.append(zrsigc[i][0])

            stresses_compr.insert(0, [i.value for i in all_rows_cs[0]][1:])
            node_i = 0
            rosette_directions = []
            for i in range(len(rosette_a_direction)):
                rosette_directions.append(rosette_a_direction[i])
                rosette_directions.append(rosette_c_direction[i])
            gauges_rosettes_a_c_indices = [i for i in range(1, R_cs - len(R_nodes) * 6)]
            # for rac in range(R_cs - len(R_nodes)*6,R_cs):
            #    gauges_rosettes_a_c_indices.append(rac)
            #    gauges_rosettes_a_c_indices.append(rac)
            for j3 in gauges_rosettes_a_c_indices:
                if ((j3 - 1) % 6 == 0):
                    node_i += 1
                stress_type = [jauge_stress_direction + rosette_directions][0][node_i - 1]
                for j4 in range(1, C_cs):
                    if (str(all_rows_cs[j3][1].value).strip() == stress_type.strip()):
                        stresses_compr[node_i].append(all_rows_cs[j3][j4].value)

            node_i = len(J_nodes) - 1
            gauges_rosettes_a_c_indices = [i for i in range(R_cs - len(R_nodes) * 6, R_cs)]
            for j3 in gauges_rosettes_a_c_indices:
                if ((j3 - 1) % 6 == 0):
                    node_i += 2
                stress_type = [jauge_stress_direction + rosette_directions][0][node_i - 1]
                for j4 in range(1, C_cs):
                    if (str(all_rows_cs[j3][1].value).strip() == stress_type.strip()):
                        stresses_compr[node_i].append(all_rows_cs[j3][j4].value)

            node_i = len(J_nodes)
            gauges_rosettes_a_c_indices = [i for i in range(R_cs - len(R_nodes) * 6, R_cs)]
            for j3 in gauges_rosettes_a_c_indices:
                if ((j3 - 1) % 6 == 0):
                    node_i += 2
                stress_type = [jauge_stress_direction + rosette_directions][0][node_i - 1]
                for j4 in range(1, C_cs):
                    if (str(all_rows_cs[j3][1].value).strip() == stress_type.strip()):
                        stresses_compr[node_i].append(all_rows_cs[j3][j4].value)


            sheet_cs_u = excel_main.create_sheet('unique_stresses_only')

            for j20 in range(len(stresses_compr[0])):
                sheet_cs_u.cell(row=1, column=j20 + 2).value = stresses_compr[0][j20]

            for j5 in range(2, len(stresses_compr) + 1):
                for j6 in range(1, len(stresses_compr[j5 - 1]) + 1):
                    sheet_cs_u.cell(row=j5, column=j6).value = stresses_compr[j5 - 1][j6 - 1]

            excel_co = openpyxl.load_workbook('correlation.xlsx')
            sheet_co = excel_co.active
            all_rows_gen_co = sheet_co.rows
            all_rows_ca = [[i for i in j]for j in all_rows_gen_co ]
            print(all_rows_ca)

            for j25 in range(0,len(J_nodes)+(len(R_nodes)*3+1)):
                for j26 in range(0,C_cs):
                    del all_rows_ca[j25][0]

            for j27 in range (len(J_nodes)+2,len(J_nodes)+(len(R_nodes)*3),3):
                for j28 in range(0,C_cs):
                    del all_rows_ca[j27][0]

            all_rows_co=[]

            for j29 in range (0,len(J_nodes)+(len(R_nodes)*3+1)):
                if all_rows_ca[j29] != []:
                    all_rows_co.append(all_rows_ca[j29])

            diff_error = [[all_rows_co[j][0].value] for j in range(2, len(all_rows_co))]
            diff_error.insert(0, all_rows_cs[0])


            diff_mat = [[all_rows_co[i][j].value - float(stresses_compr[i][j]) for j in range(2, len(stresses_compr[i]))]
                for i in range(1, (len(stresses_compr)))]

            sheet_cs_u_diff = excel_main.copy_worksheet(sheet_cs_u)
            sheet_cs_u_diff.title = 'stresses_difference'
            for j7 in range(2, len(diff_mat) + 2):
                for j8 in range(2, len(diff_mat[j7 - 2]) + 2):
                    sheet_cs_u_diff.cell(row=j7, column=j8+1).value = diff_mat[j7 - 2][j8 - 2]

            import math

            excel_input1 = openpyxl.load_workbook('correlation.xlsx')
            correlation_sheet = excel_input1['correlation']
            correlation_matrix = [[str(i.value) for i in j] for j in correlation_sheet]

            all_rows_gen_cs1 = correlation_sheet.rows
            all_columns_gen_cs1 = correlation_sheet.columns
            all_rows_cs1 = [[i for i in j] for j in all_rows_gen_cs1]
            all_columns_cs1 = [[i for i in j] for j in all_columns_gen_cs1]
            R_cs1 = len(all_rows_cs1)
            C_cs1 = len(all_columns_cs1)
            loadcase = int(C_cs1 / 2)

            j = 0
            for j1 in range(0, R_cs1):
                if correlation_matrix[j1][0][:1] == "J":
                    j += 1
            j += 1

            sheet_correlate_u = excel_input1.create_sheet('update rosettes PI PII')

            sigmastress = [correlation_matrix[i] for i in range(j, R_cs1)]

            rsigma1 = []
            rsigma2 = []
            rphi = []

            for j2 in range(0, len(sigmastress), 3):
                for j3 in range(2, loadcase):
                    r = float(sigmastress[j2][j3])
                    c = float(sigmastress[j2 + 1][j3])
                    d = float(sigmastress[j2 + 2][j3])
                    k = 0.15 * (r + d) + (0.11422 * math.sqrt((r - c) * (r - c) + (d - c) * (d - c)))
                    e = 0.15 * (r + d) - (0.11422 * math.sqrt((r - c) * (r - c) + (d - c) * (d - c)))
                    p = (math.atan2((2 * c - r - d), (r - d)))
                    p = p * 28.64
                    rsigma1.append(k)
                    rsigma2.append(e)
                    rphi.append(p)
                    sigma1 = ['%.2f' % elem for elem in rsigma1]
                    sigma2 = ['%.2f' % elem for elem in rsigma2]
                    phi = ['%.2f' % elem for elem in rphi]

            print(R_cs1)
            k = 0
            for j5 in range(1, R_cs1 + 1 - j, 3):
                for j6 in range(3, 1 + loadcase):
                    sheet_correlate_u.cell(row=j5, column=j6).value = sigma1[k]
                    k += 1
            w = 0
            for j7 in range(2, R_cs1 + 1 - j, 3):
                for j8 in range(3, 1 + loadcase):
                    sheet_correlate_u.cell(row=j7, column=j8).value = sigma2[w]
                    w += 1
            i = 0
            for j11 in range(3, R_cs1 + 1 - j, 3):
                for j12 in range(3, 1 + loadcase):
                    sheet_correlate_u.cell(row=j11, column=j12).value = phi[i]
                    i += 1

            name = []
            for j9 in range(0, len(sigmastress), 3):
                name.append(sigmastress[j9][0][:5])

            y = 0
            for j10 in range(1, R_cs1 - j, 3):
                sheet_correlate_u.cell(row=j10, column=1).value = name[y] + "sig1"
                sheet_correlate_u.cell(row=j10, column=2).value = "N/mm2"
                sheet_correlate_u.cell(row=j10 + 1, column=1).value = name[y] + "sig2"
                sheet_correlate_u.cell(row=j10 + 1, column=2).value = "N/mm2"
                sheet_correlate_u.cell(row=j10 + 2, column=1).value = name[y] + "phi"
                sheet_correlate_u.cell(row=j10 + 2, column=2).value = "grad"
                y += 1

                excel_input1.save("correlation.xlsx")


            for zone in J_nodes_zones:
                sheet_zone_name = 'Zone' + zone[-1] + '00' + zone[0].lower()
                sheet_zone = excel_main.create_sheet(sheet_zone_name)
                for j9 in range(len(stresses_compr[0])):
                    sheet_zone.cell(row=1, column=j9 + 2).value = stresses_compr[0][j9]

                row_n = 1
                for j10 in range(1, len(stresses_compr)):
                    if str(stresses_compr[j10][0][:3]) == zone and str(stresses_compr[j10][0]) in J_nodes_zones[zone]:
                        row_n += 1
                        for j11 in range(len(stresses_compr[j10])):
                            sheet_zone.cell(row=row_n, column=j11 + 1).value = stresses_compr[j10][j11]
                        # print(stresses_compr)

                row_n += 3

                r=row_n
                for j9 in range(len(all_rows_co[0])):
                    sheet_zone.cell(row=row_n, column=j9 + 1).value = all_rows_co[0][j9].value
                for j10 in range(1, len(all_rows_co)):
                    if str(all_rows_co[j10][0].value[:3]) == zone and str(all_rows_co[j10][0].value) in J_nodes_zones[zone]:
                        row_n += 1
                        for j11 in range(len(all_rows_co[j10])):
                            sheet_zone.cell(row=row_n, column=j11 + 1).value = all_rows_co[j10][j11].value
                s=row_n+1
                k=1
                for z in range(r,s):
                    sheet_zone.cell(row=z, column=2).value=sheet_zone.cell(row=k, column= 2).value
                    k +=1

                row_n += 3
                d=row_n
                for j9 in range(len(all_rows_co[0])):
                    sheet_zone.cell(row=row_n, column=j9 + 1).value = all_rows_co[0][j9].value
                for j10 in range(1, len(diff_mat) + 1):
                    if str(all_rows_co[j10][0].value[:3]) == zone and str(all_rows_co[j10][0].value) in J_nodes_zones[zone]:

                        row_n += 1
                        sheet_zone.cell(row=row_n, column=1).value = all_rows_co[j10][0].value
                        for j11 in range(1, len(diff_mat[j10 - 1]) + 1):
                            sheet_zone.cell(row=row_n, column=j11 + 2).value = diff_mat[j10 - 1][j11 - 1]
                e=row_n+1
                q=1
                for z in range(d,e):
                    sheet_zone.cell(row=z, column=2).value = sheet_zone.cell(row=q, column=2).value
                    q += 1




            for zone in R_nodes_zones:
                sheet_zone_name = 'Zone' + zone[-1] + '00' + zone[0].lower()
                sheet_zone = excel_main.create_sheet(sheet_zone_name)
                for j9 in range(len(stresses_compr[0])):
                    sheet_zone.cell(row=1, column=j9 + 2).value = stresses_compr[0][j9]

                row_n = 1
                for j10 in range(1, len(stresses_compr)):
                    if str(stresses_compr[j10][0][:3]) == zone and str(stresses_compr[j10][0][:5]) in R_nodes_zones[zone]:

                        row_n += 1
                        for j11 in range(len(stresses_compr[j10])):
                            sheet_zone.cell(row=row_n, column=j11 + 1).value = stresses_compr[j10][j11]


                row_n += 3
                l=row_n

                for j9 in range(len(all_rows_co[0])):
                    sheet_zone.cell(row=row_n, column=j9 + 1).value = all_rows_co[0][j9].value
                for j10 in range(1, len(all_rows_co)):
                    if str(all_rows_co[j10][0].value[:3]) == zone and str(all_rows_co[j10][0].value[:5]) in R_nodes_zones[zone]:

                        row_n += 1
                        for j11 in range(len(all_rows_co[j10])):
                            sheet_zone.cell(row=row_n, column=j11 + 1).value = all_rows_co[j10][j11].value
                w=row_n+1
                u=1
                for z in range(l,w):
                    sheet_zone.cell(row=z, column=2).value = sheet_zone.cell(row=u, column=2).value
                    u += 1


                row_n += 3
                f=row_n
                for j9 in range(len(all_rows_co[0])):
                    sheet_zone.cell(row=row_n, column=j9 + 1).value = all_rows_co[0][j9].value
                for j10 in range(1, len(diff_mat) + 1):
                    if str(all_rows_co[j10][0].value[:3]) == zone and str(all_rows_co[j10][0].value[:5]) in R_nodes_zones[zone]:
                        row_n += 1
                        sheet_zone.cell(row=row_n, column=1).value = all_rows_co[j10][0].value
                        for j11 in range(1, len(diff_mat[j10 - 1]) + 1):
                            sheet_zone.cell(row=row_n, column=j11 + 2).value = diff_mat[j10 - 1][j11 - 1]
                c=row_n+1
                b=1
                for z in range(f,c):
                    sheet_zone.cell(row=z, column=2).value = sheet_zone.cell(row=b, column=2).value
                    b += 1
                excel_main.save("stress_sheets_final.xlsx")