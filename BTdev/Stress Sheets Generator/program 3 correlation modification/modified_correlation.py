#created by dhruvin jasoliya


import openpyxl
import math

excel_input1= openpyxl.load_workbook('correlation.xlsx')
correlation_sheet=excel_input1['correlation']
correlation_matrix=[[str(i.value)for i in j]for j in correlation_sheet]

all_rows_gen_cs1 = correlation_sheet.rows
all_columns_gen_cs1 = correlation_sheet.columns
all_rows_cs1 = [[i for i in j] for j in all_rows_gen_cs1]
all_columns_cs1 = [[i for i in j] for j in all_columns_gen_cs1]
R_cs1 = len(all_rows_cs1)
C_cs1 = len(all_columns_cs1)
loadcase=int(C_cs1/2)

j=0
for j1 in range (0,R_cs1):
    if correlation_matrix[j1][0][:1]== "J" :
        j +=1
j += 1

sheet_correlate_u=excel_input1.create_sheet('update rosettes PI PII')


sigmastress=[correlation_matrix[i]for i in range(j,R_cs1) ]

rsigma1=[]
rsigma2=[]
rphi=[]

for j2 in range (0,len(sigmastress),3):
    for j3 in range (2,loadcase):
        r=float(sigmastress[j2][j3])
        c=float(sigmastress[j2+1][j3])
        d=float(sigmastress[j2+2][j3])
        k=0.15*(r+d)+(0.11422*math.sqrt((r-c)*(r-c)+(d-c)*(d-c)))
        e=0.15*(r+d)-(0.11422*math.sqrt((r-c)*(r-c)+(d-c)*(d-c)))
        p=(math.atan2((2*c-r-d),(r-d)))
        p=p*28.64
        rsigma1.append(k)
        rsigma2.append(e)
        rphi.append(p)
        sigma1 = ['%.2f' % elem for elem in rsigma1]
        sigma2 = ['%.2f' % elem for elem in rsigma2]
        phi= ['%.2f' % elem for elem in rphi]

print(R_cs1)
k=0
for j5 in range (1,R_cs1+1-j,3):
    for j6 in range(3,1+loadcase):
        sheet_correlate_u.cell(row=j5, column=j6).value = sigma1[k]
        k+=1
w=0
for j7 in range (2,R_cs1+1-j,3):
    for j8 in range(3,1+loadcase):
        sheet_correlate_u.cell(row=j7, column=j8).value = sigma2[w]
        w+=1
i=0
for j11 in range (3,R_cs1+1-j,3):
    for j12 in range(3,1+loadcase):
        sheet_correlate_u.cell(row=j11, column=j12).value = phi[i]
        i+=1


name=[]
for j9 in range (0,len(sigmastress),3):
    name.append(sigmastress[j9][0][:5])

y=0
for j10 in range(1,R_cs1-j,3):
    sheet_correlate_u.cell(row=j10, column=1).value = name[y]+ "sig1"
    sheet_correlate_u.cell(row=j10, column= 2).value = "N/mm2"
    sheet_correlate_u.cell(row=j10+1, column=1).value =name[y]+ "sig2"
    sheet_correlate_u.cell(row=j10 + 1, column= 2).value =  "N/mm2"
    sheet_correlate_u.cell(row=j10+2, column=1).value = name[y]+"phi"
    sheet_correlate_u.cell(row=j10 + 2, column= 2).value = "grad"
    y +=1

    excel_input1.save("correlation.xlsx")












